"""Agreement-only IRR report from precomputed qN.agree boolean columns.

This script complements `compute_irr.py`. The full pipeline (`compute_irr.py`)
reconstructs Cohen / Fleiss / Krippendorff / Jaccard from per-coder answer
cells; this one trusts the workbook's precomputed `qN.agree` flags and
reports just the raw % agreement per question.

When to use which:
- `compute_irr.py`: rigorous IRR. Drops "both coders left blank" rows and
  computes chance-corrected kappa/alpha. Surfaces sparsity in the per-coder
  data.
- `compute_agreement_only.py` (this file): matches how the source workbook
  itself reports agreement (qN.agree True/False per row). Useful as the
  "official" first-pass agreement number, but cannot report kappa/alpha
  because the per-coder answers aren't read.

Usage:
    /home/G39248410/citizen_voice/venv/bin/python compute_agreement_only.py \\
        --input "input/2025-06-30 Second Coding Check _ Volume 134 - Part 2.xlsx" \\
        --sheet "104 agreement" \\
        --output-xlsx outputs/agreement_only_vol134p2_104.xlsx \\
        --output-md outputs/agreement_only_vol134p2_104.md
"""
from __future__ import annotations

import argparse
import math
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl


QUESTIONS = ("q1", "q2", "q3", "q4", "q5", "q6")
CODERS = ("Leah", "Bridget", "Rachel", "Alia")


def load_agreement_sheet(path: Path, sheet: str) -> Dict:
    """Load qN.agree booleans (and q1.rest.<coder> details) from the named sheet.

    Returns:
        {
            "source": Path,
            "sheet": str,
            "n_rows": int,
            "per_question": {
                "q1": {
                    "n": int,                # rows where the flag is True or False
                    "n_agree": int,
                    "n_disagree": int,
                    "pct_agreement": float,  # n_agree / n (NaN if n == 0)
                    "disagreement_rows": [{"order_num": int, ...}, ...],
                },
                ...
            },
            "coders_listed": [str, ...],  # coders named in the q1.rest.* columns
        }
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(
            f"sheet {sheet!r} not in workbook. Available: {wb.sheetnames}"
        )
    ws = wb[sheet]
    header = list(next(ws.iter_rows(values_only=True)))
    col_idx = {h: i for i, h in enumerate(header)}

    # Required columns
    if "order.num" not in col_idx:
        raise ValueError("sheet missing required column 'order.num'")
    for q in QUESTIONS:
        if f"{q}.agree" not in col_idx:
            raise ValueError(f"sheet missing required column {q}.agree")

    # Optional: q1.rest.<coder> columns enrich Q1 disagreement listings
    q1_rest_cols = {
        c: col_idx[f"q1.rest.{c}"]
        for c in CODERS
        if f"q1.rest.{c}" in col_idx
    }

    per_question = {q: {"n": 0, "n_agree": 0, "n_disagree": 0, "disagreement_rows": []}
                    for q in QUESTIONS}
    n_rows = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or row[col_idx["order.num"]] is None:
            continue
        n_rows += 1
        order_num = int(row[col_idx["order.num"]])
        for q in QUESTIONS:
            v = row[col_idx[f"{q}.agree"]]
            if v is True:
                per_question[q]["n_agree"] += 1
                per_question[q]["n"] += 1
            elif v is False:
                per_question[q]["n_disagree"] += 1
                per_question[q]["n"] += 1
                # Record disagreement
                drow = {"order_num": order_num}
                if q == "q1":
                    for c, ci in q1_rest_cols.items():
                        drow[c] = row[ci]
                per_question[q]["disagreement_rows"].append(drow)
            # else: None / blank → not counted in denominator

    for q in QUESTIONS:
        n = per_question[q]["n"]
        per_question[q]["pct_agreement"] = (
            per_question[q]["n_agree"] / n if n > 0 else math.nan
        )

    return {
        "source": Path(path),
        "sheet": sheet,
        "n_rows": n_rows,
        "per_question": per_question,
        "coders_listed": list(q1_rest_cols.keys()),
    }


def _ceiling_band(pct: float) -> str:
    """Map raw % agreement to a qualitative ceiling label.

    These bands are stricter than the Landis-Koch bands for kappa because
    raw % is uncorrected for chance — a 95% agreement on a binary question
    where 95% of answers are 'no' is unimpressive.
    """
    if pct is None or (isinstance(pct, float) and math.isnan(pct)):
        return "n/a"
    if pct < 0.70:
        return "poor"
    if pct < 0.85:
        return "moderate"
    if pct < 0.95:
        return "substantial"
    return "almost perfect"


def _fmt_pct(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return "NaN"
        return f"{value:.4f}"
    return str(value)


def write_excel(result: Dict, out_path: Path) -> None:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "summary"

    summary.append([
        "question", "n", "n_agree", "n_disagree", "pct_agreement", "ceiling_band",
    ])
    for q in QUESTIONS:
        m = result["per_question"][q]
        summary.append([
            q.upper(),
            m["n"],
            m["n_agree"],
            m["n_disagree"],
            _fmt_pct(m["pct_agreement"]),
            _ceiling_band(m["pct_agreement"]),
        ])

    dis = wb.create_sheet("disagreements")
    coders = result["coders_listed"]
    dis.append(["question", "order_num"] + coders)
    for q in QUESTIONS:
        for drow in result["per_question"][q]["disagreement_rows"]:
            dis.append(
                [q.upper(), drow["order_num"]]
                + [drow.get(c, "") if drow.get(c) is not None else "" for c in coders]
            )

    meta = wb.create_sheet("metadata")
    meta.append(["key", "value"])
    meta.append(["source", str(result["source"])])
    meta.append(["sheet", result["sheet"]])
    meta.append(["n_rows", result["n_rows"]])
    meta.append(["coders_listed_in_q1_rest", ", ".join(coders) if coders else "(none)"])
    meta.append(["generated_at_utc", datetime.now(timezone.utc).isoformat(timespec="seconds")])
    meta.append([
        "methodology",
        "Raw % agreement from precomputed qN.agree boolean columns. "
        "kappa / alpha / Jaccard not computed (require per-coder responses).",
    ])
    meta.append([
        "ceiling_bands",
        "%<0.70 poor; 0.70-0.85 moderate; 0.85-0.95 substantial; >=0.95 almost perfect",
    ])

    wb.save(out_path)


def write_markdown(result: Dict, out_path: Path) -> None:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    src = result["source"].name
    coders = ", ".join(result["coders_listed"]) if result["coders_listed"] else "(none in q1.rest)"

    lines = []
    lines.append("# Inter-Rater Agreement Report (boolean-only)")
    lines.append("")
    lines.append(f"- **Source:** `{src}`")
    lines.append(f"- **Sheet:** `{result['sheet']}`")
    lines.append(f"- **Rows in sheet:** {result['n_rows']}")
    lines.append(f"- **Coders in q1.rest columns:** {coders}")
    lines.append(f"- **Generated (UTC):** {datetime.now(timezone.utc).isoformat(timespec='seconds')}")
    lines.append("")
    lines.append("## Per-question agreement")
    lines.append("")
    lines.append("| Question | N | # agree | # disagree | % agreement | Ceiling |")
    lines.append("|---|---|---|---|---|---|")
    for q in QUESTIONS:
        m = result["per_question"][q]
        lines.append(
            f"| {q.upper()} | {m['n']} | {m['n_agree']} | {m['n_disagree']} | "
            f"{_fmt_pct(m['pct_agreement'])} | {_ceiling_band(m['pct_agreement'])} |"
        )
    lines.append("")
    lines.append("## Disagreements")
    lines.append("")
    any_dis = False
    for q in QUESTIONS:
        rows = result["per_question"][q]["disagreement_rows"]
        if not rows:
            continue
        any_dis = True
        lines.append(f"### {q.upper()} ({len(rows)} row{'s' if len(rows)!=1 else ''})")
        lines.append("")
        if q == "q1" and result["coders_listed"]:
            lines.append("| order_num | " + " | ".join(result["coders_listed"]) + " |")
            lines.append("|---|" + "|".join("---" for _ in result["coders_listed"]) + "|")
            for drow in rows:
                cells = [drow.get(c, "") if drow.get(c) is not None else "" for c in result["coders_listed"]]
                lines.append(f"| {drow['order_num']} | " + " | ".join(str(x) for x in cells) + " |")
        else:
            lines.append("| order_num |")
            lines.append("|---|")
            for drow in rows:
                lines.append(f"| {drow['order_num']} |")
        lines.append("")
    if not any_dis:
        lines.append("(none — every row marked agree=True)")
        lines.append("")

    lines.append("## Methodology")
    lines.append("")
    lines.append(
        "Agreement is taken directly from the workbook's precomputed `qN.agree` "
        "columns in the specified sheet. The denominator is the number of rows "
        "where the flag is True or False; rows where the flag is blank (None) are "
        "excluded. **No per-coder answer cells are read**, so chance-corrected "
        "metrics (Cohen's κ, Fleiss' κ, Krippendorff's α) and label-set agreement "
        "(Jaccard) are not computed. Ceiling bands (poor / moderate / substantial / "
        "almost perfect) use stricter thresholds than the Landis-Koch κ bands "
        "because raw % agreement is uncorrected for chance."
    )
    lines.append("")
    lines.append("## Caveats")
    lines.append("")
    lines.append(
        "- The `qN.agree` columns may treat *'both coders left this blank'* as "
        "agreement. On conditional questions (Q2–Q5, only filled when Q1=yes), "
        "this inflates % agreement relative to a per-coder reconstruction."
    )
    lines.append(
        "- For Q1 disagreements, the `q1.rest.<coder>` columns show each coder's "
        "Q1 answer side-by-side. For Q2–Q6, only the disagreement order numbers "
        "are listed (the source sheet has no per-coder detail for those)."
    )
    lines.append(
        "- For chance-corrected metrics and label-set agreement, run "
        "`compute_irr.py` against the same workbook (it reads the per-coder Q.* "
        "cells from the coder sheets instead)."
    )
    lines.append("")

    out_path.write_text("\n".join(lines))


def main():
    parser = argparse.ArgumentParser(
        description="Compute raw % agreement per question from the precomputed qN.agree columns."
    )
    parser.add_argument("--input", required=True, type=Path,
                        help="Path to a workbook containing an agreement sheet")
    parser.add_argument("--sheet", required=True,
                        help="Name of the agreement sheet (e.g. '104 agreement')")
    parser.add_argument("--output-xlsx", required=True, type=Path)
    parser.add_argument("--output-md", required=True, type=Path)
    args = parser.parse_args()

    result = load_agreement_sheet(args.input, args.sheet)
    write_excel(result, args.output_xlsx)
    write_markdown(result, args.output_md)
    print(f"wrote {args.output_xlsx}")
    print(f"wrote {args.output_md}")


if __name__ == "__main__":
    main()
