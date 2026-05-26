"""Load a citizen_voice cross_q_check workbook into long-form.

The cross_q_check format is described in the IRR pilot spec, §3. Briefly:
one sheet, one row per `order.num`, with per-coder per-question columns
named `Q1.Leah`, `Q1.notes.Leah`, ..., `Q6.Alia`, `notes.Alia`. Coders that
did not code a given row leave their cells blank (None).

This loader returns a long-form DataFrame with columns:
    order_num: int
    coder:     str (one of the 4 known coder names)
    question:  str ("Q1".."Q6")
    answer:    str or None (None == blank or "(blank)")
    notes:     str or None

Only `(order_num, coder, question)` rows where the coder *participated*
in that order are emitted. Participation is determined by the coder
indicator columns (the bare `Leah`, `Bridget`, `Rachel`, `Alia` columns
with "1" if participated, blank otherwise).
"""
from pathlib import Path
from typing import Union

import openpyxl
import pandas as pd

CODERS = ("Leah", "Bridget", "Rachel", "Alia")
QUESTIONS = ("Q1", "Q2", "Q3", "Q4", "Q5", "Q6")


def _blank_to_none(value):
    """Normalize the various ways the source file represents 'blank'."""
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == "" or stripped == "(blank)":
            return None
        return stripped
    return value


def load_cross_q_check(path: Union[str, Path]) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    header = next(ws.iter_rows(values_only=True))
    col_idx = {h: i for i, h in enumerate(header)}

    # Per-coder participation indicator: the bare column named "Leah", "Bridget", etc.
    indicator_cols = {c: col_idx[c] for c in CODERS if c in col_idx}
    if len(indicator_cols) != len(CODERS):
        missing = set(CODERS) - set(indicator_cols)
        raise ValueError(f"cross_q_check schema missing coder indicator columns: {missing}")

    # For notes: Q1..Q5 have `Q<n>.notes.<coder>`; Q6 has `notes.<coder>`.
    def notes_col(coder, q):
        if q == "Q6":
            return f"notes.{coder}"
        return f"Q{q[1]}.notes.{coder}"

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or row[0] is None:
            continue
        order_num = int(row[col_idx["order.num"]])
        for coder, ind_col in indicator_cols.items():
            participated = _blank_to_none(row[ind_col]) is not None
            if not participated:
                continue
            for q in QUESTIONS:
                ans_col = f"{q}.{coder}"
                if ans_col not in col_idx:
                    continue
                ans = _blank_to_none(row[col_idx[ans_col]])
                ncol = notes_col(coder, q)
                note = _blank_to_none(row[col_idx[ncol]]) if ncol in col_idx else None
                records.append({
                    "order_num": order_num,
                    "coder": coder,
                    "question": q,
                    "answer": ans,
                    "notes": note,
                })

    # pandas 3.x infers "str" dtype for string-like columns and converts
    # Python None -> NaN.  Build answer/notes as explicit object-dtype Series
    # so that callers can reliably use `value is None` checks.
    import numpy as np

    n = len(records)
    ans_arr = np.empty(n, dtype=object)
    notes_arr = np.empty(n, dtype=object)
    for i, r in enumerate(records):
        ans_arr[i] = r["answer"]
        notes_arr[i] = r["notes"]

    df = pd.DataFrame({
        "order_num": [r["order_num"] for r in records],
        "coder": [r["coder"] for r in records],
        "question": [r["question"] for r in records],
        "answer": pd.Series(ans_arr, dtype=object),
        "notes": pd.Series(notes_arr, dtype=object),
    })[["order_num", "coder", "question", "answer", "notes"]]
    return df
