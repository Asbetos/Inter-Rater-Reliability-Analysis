"""Build Excel + Markdown reports from the dict returned by compute_irr.run."""
from __future__ import annotations

import math
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict

import openpyxl


def _ceiling_band(alpha: float) -> str:
    """Map Krippendorff's alpha to a 4-band qualitative ceiling label."""
    if alpha is None or (isinstance(alpha, float) and math.isnan(alpha)):
        return "n/a"
    if alpha < 0.4:
        return "poor"
    if alpha < 0.6:
        return "moderate"
    if alpha < 0.8:
        return "substantial"
    return "almost perfect"


def _fmt(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return "NaN"
        return f"{value:.4f}"
    return str(value)


def write_excel(result: Dict, out_path: Path) -> None:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    # Default sheet is named 'Sheet'; we rename rather than create+remove.
    summary = wb.active
    summary.title = "summary"

    summary.append([
        "question", "type", "n_units", "raw_pct_agreement",
        "mean_pairwise_cohen", "krippendorff_alpha", "fleiss_kappa",
        "strict_pct_agreement", "mean_jaccard",
        "ceiling_band", "underpowered",
    ])
    for q, m in result["per_question"].items():
        summary.append([
            q, m["type"], m["n_units"],
            _fmt(m["raw_pct_agreement"]),
            _fmt(m["mean_pairwise_cohen"]),
            _fmt(m["krippendorff_alpha"]),
            _fmt(m["fleiss_kappa"]),
            _fmt(m["strict_pct_agreement"]),
            _fmt(m["mean_jaccard"]),
            _ceiling_band(m["krippendorff_alpha"]),
            "yes" if m["underpowered"] else "no",
        ])

    # Pairwise kappa sheet
    pk = wb.create_sheet("pairwise_kappa")
    pk.append(["question", "coder_a", "coder_b", "cohen_kappa"])
    for q, m in result["per_question"].items():
        for (a, b), k in m["pairwise_cohen"].items():
            pk.append([q, a, b, _fmt(k)])

    # Disagreements sheet
    dis = wb.create_sheet("disagreements")
    coders = result["coders_present"]
    dis.append(["question", "order_num"] + coders)
    for q, m in result["per_question"].items():
        for d in m["disagreements"]:
            dis.append([q, d["order_num"]] + [d.get(c, "") if d.get(c) is not None else "" for c in coders])

    # Metadata sheet
    meta = wb.create_sheet("metadata")
    meta.append(["key", "value"])
    meta.append(["source", str(result["source"])])
    meta.append(["n_orders", result["n_orders"]])
    meta.append(["coders_present", ", ".join(result["coders_present"])])
    meta.append(["generated_at_utc", datetime.now(timezone.utc).isoformat(timespec="seconds")])
    meta.append(["ceiling_bands", "alpha<0.4 poor; 0.4-0.6 moderate; 0.6-0.8 substantial; >=0.8 almost perfect"])

    wb.save(out_path)


def write_markdown(result: Dict, out_path: Path) -> None:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    src = result["source"].name if isinstance(result["source"], Path) else str(result["source"])
    coders = ", ".join(result["coders_present"])

    lines = []
    lines.append("# Inter-Rater Reliability Report")
    lines.append("")
    lines.append(f"- **Source:** `{src}`")
    lines.append(f"- **Orders in workbook:** {result['n_orders']}")
    lines.append(f"- **Coders present:** {coders}")
    lines.append(f"- **Generated (UTC):** {datetime.now(timezone.utc).isoformat(timespec='seconds')}")
    lines.append("")
    lines.append("## Per-question summary")
    lines.append("")
    lines.append("| Question | Type | N (units) | Raw % agree | Mean Cohen κ | Krippendorff α | Ceiling | Underpowered? |")
    lines.append("|---|---|---|---|---|---|---|---|")
    for q, m in result["per_question"].items():
        lines.append(
            f"| {q} | {m['type']} | {m['n_units']} | "
            f"{_fmt(m['raw_pct_agreement'])} | "
            f"{_fmt(m['mean_pairwise_cohen'])} | "
            f"{_fmt(m['krippendorff_alpha'])} | "
            f"{_ceiling_band(m['krippendorff_alpha'])} | "
            f"{'yes' if m['underpowered'] else 'no'} |"
        )
    lines.append("")
    lines.append("## Per-question detail")
    for q, m in result["per_question"].items():
        lines.append("")
        lines.append(f"### {q} — {m['type']}")
        lines.append("")
        if m["type"] in ("multi_label", "compound_categorical"):
            lines.append(f"- Strict % agreement: {_fmt(m['strict_pct_agreement'])}")
            lines.append(f"- Mean Jaccard: {_fmt(m['mean_jaccard'])}")
        lines.append(f"- N units: {m['n_units']} (underpowered if <30: {'yes' if m['underpowered'] else 'no'})")
        lines.append(f"- Raw % agreement: {_fmt(m['raw_pct_agreement'])}")
        if m["pairwise_cohen"]:
            lines.append("- Pairwise Cohen κ:")
            for (a, b), k in m["pairwise_cohen"].items():
                lines.append(f"  - {a} vs {b}: {_fmt(k)}")
        lines.append(f"- Krippendorff α: {_fmt(m['krippendorff_alpha'])}")
        if not (isinstance(m["fleiss_kappa"], float) and math.isnan(m["fleiss_kappa"])):
            lines.append(f"- Fleiss κ (rows with all coders): {_fmt(m['fleiss_kappa'])}")
        if m["disagreements"]:
            lines.append(f"- Disagreement rows ({len(m['disagreements'])}): see `disagreements` sheet")

    lines.append("")
    lines.append("## Methodology")
    lines.append("")
    lines.append(
        "Inter-rater reliability is computed from the per-coder answer columns "
        "of the cross_q_check workbook, not from the precomputed `qN.agree` flags. "
        "For binary questions (Q1, Q2, Q3, Q6), units are included where at least "
        "two coders gave non-blank answers. For multi-label / compound questions "
        "(Q4, Q5), units are included where at least one coder gave a non-empty "
        "answer set; both strict (exact-match) and lenient (Jaccard) agreement "
        "are reported. The Krippendorff α and ceiling-band columns are the "
        "recommended summary statistics: bands are < 0.4 poor; 0.4–0.6 moderate; "
        "0.6–0.8 substantial; ≥ 0.8 almost perfect (Landis & Koch 1977, applied to α)."
    )
    lines.append("")
    lines.append("## Caveats")
    lines.append("")
    lines.append("- Pilot on a single volume and single (earliest) snapshot.")
    lines.append("- Q2/Q3/Q4/Q5 are conditional on Q1=yes; their N is small and the underpowered flag is set if N < 30.")
    lines.append("- Multi-label Q4 has two views (strict and Jaccard); choose the one that matches your LLM-benchmark scoring rule.")
    lines.append("- The earliest snapshot may already include partial reconciliation; verify against any older raw-coder exports if available.")
    lines.append("")

    out_path.write_text("\n".join(lines))
