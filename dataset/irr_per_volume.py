"""Compute the three IRR variants for one volume.

Reads per-coder sheets directly from the workbook (sheets named "Leah",
"Bridget", "Rachel", "Alia", "Brian"). Returns two lists of records:
one per (coder, question) and one per (coder_pair, question).

Reuses the existing pilot's primitives:
    - irr_metrics.cohen_kappa
    - normalize.normalize_categorical, normalize.split_label_set
"""
from __future__ import annotations

import math
from collections import Counter
from itertools import combinations
from typing import Dict, List, Tuple

import openpyxl

import irr_metrics
import normalize


KNOWN_CODERS = ("Leah", "Bridget", "Rachel", "Alia", "Brian")
QUESTIONS_DEFAULT = ("Q1", "Q2", "Q3", "Q4", "Q5", "Q6")
MULTI_LABEL_QUESTIONS = ("Q4",)
COMPOUND_CATEGORICAL_QUESTIONS = ("Q5",)


def _detect_coders(wb: openpyxl.Workbook) -> List[str]:
    return [c for c in KNOWN_CODERS if c in wb.sheetnames]


def _detect_questions(wb: openpyxl.Workbook, coders: List[str]) -> List[str]:
    """Read the first coder sheet's header to detect the Q1..Q? range."""
    if not coders:
        return list(QUESTIONS_DEFAULT)
    ws = wb[coders[0]]
    header = next(ws.iter_rows(values_only=True), None)
    if header is None:
        return list(QUESTIONS_DEFAULT)
    q_seen = set()
    for h in header:
        if not (isinstance(h, str) and h):
            continue
        # Match "1." or "Q1." or "1. (..."
        stripped = h.strip()
        if "." in stripped:
            prefix = stripped.split(".", 1)[0].strip()
            if prefix.upper().startswith("Q"):
                prefix = prefix[1:]
            if prefix.isdigit():
                q_seen.add(f"Q{prefix}")
    return sorted(q_seen) if q_seen else list(QUESTIONS_DEFAULT)


def _load_per_coder_answers(
    wb: openpyxl.Workbook, coder: str, questions: List[str]
) -> Dict[int, Dict[str, object]]:
    """Return {order_num: {question: normalized_answer}} for one coder's sheet."""
    ws = wb[coder]
    header = list(next(ws.iter_rows(values_only=True), []))
    q_col_idx: Dict[str, int] = {}
    for i, h in enumerate(header):
        if not (isinstance(h, str) and h):
            continue
        stripped = h.strip()
        if "." not in stripped:
            continue
        prefix = stripped.split(".", 1)[0].strip()
        if prefix.upper().startswith("Q"):
            prefix = prefix[1:]
        if not prefix.isdigit():
            continue
        q_key = f"Q{prefix}"
        if q_key in questions and q_key not in q_col_idx:
            q_col_idx[q_key] = i

    result: Dict[int, Dict[str, object]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or row[0] is None:
            continue
        try:
            order = int(row[0])
        except (TypeError, ValueError):
            continue
        per_q = {}
        for q, idx in q_col_idx.items():
            raw = row[idx] if idx < len(row) else None
            if q in MULTI_LABEL_QUESTIONS or q in COMPOUND_CATEGORICAL_QUESTIONS:
                per_q[q] = normalize.split_label_set(raw)
            else:
                per_q[q] = normalize.normalize_categorical(raw)
        result[order] = per_q
    return result


def _is_blank_value(v, question: str) -> bool:
    if question in MULTI_LABEL_QUESTIONS or question in COMPOUND_CATEGORICAL_QUESTIONS:
        return v is None or (hasattr(v, "__len__") and len(v) == 0)
    return v is None


def _to_kappa_token(v, question: str):
    """Convert a frozenset to a canonical string for Cohen kappa; pass through otherwise."""
    if question in MULTI_LABEL_QUESTIONS or question in COMPOUND_CATEGORICAL_QUESTIONS:
        if v is None or len(v) == 0:
            return None
        return ",".join(sorted(v))
    return v


def compute(wb: openpyxl.Workbook) -> Tuple[List[Dict], List[Dict]]:
    coders = _detect_coders(wb)
    questions = _detect_questions(wb, coders)
    answers = {c: _load_per_coder_answers(wb, c, questions) for c in coders}

    per_coder_rows: List[Dict] = []
    per_pair_rows: List[Dict] = []

    for q in questions:
        # Build {order: {coder: value}} restricted to non-blank values.
        order_to_coder_val: Dict[int, Dict[str, object]] = {}
        for c in coders:
            for order, per_q in answers[c].items():
                v = per_q.get(q)
                if not _is_blank_value(v, q):
                    order_to_coder_val.setdefault(order, {})[c] = v

        # Variant A: leave-one-out raw % agreement (per coder)
        # Variant B: mean pairwise Cohen kappa (per coder)
        for i in coders:
            n_orders_eligible = 0
            agreed = 0
            for order, cv in order_to_coder_val.items():
                if i not in cv:
                    continue
                others = {c: v for c, v in cv.items() if c != i}
                if not others:
                    continue
                tokens = [_to_kappa_token(v, q) for v in others.values()]
                modal_count = Counter(tokens).most_common(1)[0][1]
                tied = [t for t, c in Counter(tokens).items() if c == modal_count]
                # Deterministic tie-break: lexicographic min (Nones sort last).
                modal_token = sorted(tied, key=lambda x: (x is None, str(x)))[0]
                my_token = _to_kappa_token(cv[i], q)
                n_orders_eligible += 1
                if my_token == modal_token:
                    agreed += 1
            irr_loo = agreed / n_orders_eligible if n_orders_eligible else float("nan")

            # Variant B
            kappa_values = []
            for j in coders:
                if j == i:
                    continue
                xi, xj = [], []
                for order, cv in order_to_coder_val.items():
                    if i in cv and j in cv:
                        xi.append(_to_kappa_token(cv[i], q))
                        xj.append(_to_kappa_token(cv[j], q))
                if not xi:
                    continue
                k = irr_metrics.cohen_kappa(xi, xj)
                if not math.isnan(k):
                    kappa_values.append(k)
            irr_mean_kappa = (
                sum(kappa_values) / len(kappa_values) if kappa_values else float("nan")
            )

            n_disagreements = (
                n_orders_eligible - agreed if not math.isnan(irr_loo) else 0
            )
            per_coder_rows.append({
                "coder": i,
                "question": q,
                "irr_leave_one_out": irr_loo,
                "irr_mean_pairwise_kappa": irr_mean_kappa,
                "n_orders_eligible": n_orders_eligible,
                "n_disagreements": n_disagreements,
            })

        # Variant C: per-pair kappa
        for a, b in combinations(sorted(coders), 2):
            xa, xb = [], []
            for order, cv in order_to_coder_val.items():
                if a in cv and b in cv:
                    xa.append(_to_kappa_token(cv[a], q))
                    xb.append(_to_kappa_token(cv[b], q))
            n_overlap = len(xa)
            n_dis = sum(1 for x, y in zip(xa, xb) if x != y)
            kappa = irr_metrics.cohen_kappa(xa, xb) if n_overlap else float("nan")
            per_pair_rows.append({
                "coder_a": a,
                "coder_b": b,
                "question": q,
                "irr_pairwise_kappa": kappa,
                "n_overlap_orders": n_overlap,
                "n_disagreements": n_dis,
            })

    return per_coder_rows, per_pair_rows
