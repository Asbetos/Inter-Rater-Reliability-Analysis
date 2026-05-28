import openpyxl

import tracking_oracle


def test_tracking_oracle_empty_dict_when_tracker_missing(tmp_path):
    """If the DriveSource has no tracker file, the oracle returns {}."""
    import drive_source
    # Create a directory with one xlsx that does NOT match the pattern
    (tmp_path / "Volume-Z.xlsx").write_bytes(b"")  # won't even need to open it
    src = drive_source.LocalDriveSource(tmp_path)
    result = tracking_oracle.load(src, tracker_filename_pattern="tracking status")
    assert result == {}


def test_tracking_oracle_parses_simple_mock_tracker(tmp_path):
    """Build a one-off mock tracker xlsx and confirm parsing."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Schema: volume | status | leah | bridget | rachel | alia
    ws.append(["volume", "status", "Leah", "Bridget", "Rachel", "Alia"])
    ws.append(["134_p2", "complete", "yes", "yes", "yes", "yes"])
    ws.append(["135",     "in progress", "yes", "no", "yes", "no"])
    wb.save(tmp_path / "tracking_status_mock.xlsx")
    import drive_source
    src = drive_source.LocalDriveSource(tmp_path)
    result = tracking_oracle.load(src, tracker_filename_pattern="tracking_status")
    assert "134_p2" in result
    assert result["134_p2"]["complete"] is True
    assert set(result["134_p2"]["expected_coders"]) == {"Leah", "Bridget", "Rachel", "Alia"}
    assert "135" in result
    assert result["135"]["complete"] is False
    assert set(result["135"]["expected_coders"]) == {"Leah", "Rachel"}


def test_tracking_oracle_returns_empty_on_missing_volume_column(tmp_path):
    """Tracker without a 'volume' column is unrecognized, return {} (tolerant)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["some_other_column", "status"])
    ws.append(["x", "complete"])
    wb.save(tmp_path / "tracking_status_weird.xlsx")
    import drive_source
    src = drive_source.LocalDriveSource(tmp_path)
    result = tracking_oracle.load(src, tracker_filename_pattern="tracking_status")
    assert result == {}


def test_tracking_oracle_handles_alternate_status_tokens(tmp_path):
    """Aliases for 'complete' must work, and non-aliases default to False."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["volume", "status"])
    ws.append(["A", "DONE"])
    ws.append(["B", "Finished"])
    ws.append(["C", "y"])   # NOT in alias set ("yes" is the token, not "y")  -> False
    ws.append(["D", "complete"])
    ws.append(["E", "pending"])
    wb.save(tmp_path / "tracking_status_aliases.xlsx")
    import drive_source
    src = drive_source.LocalDriveSource(tmp_path)
    result = tracking_oracle.load(src, tracker_filename_pattern="tracking_status")
    assert result["A"]["complete"] is True   # "done"
    assert result["B"]["complete"] is True   # "finished"
    assert result["C"]["complete"] is False  # "y" is for coder columns, not status
    assert result["D"]["complete"] is True   # "complete"
    assert result["E"]["complete"] is False  # "pending"
