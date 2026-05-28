"""Build the mini-Drive fixture (4 mock workbooks).

Run from irr_analysis/:
    /home/G39248410/citizen_voice/venv/bin/python dataset/tests/build_mini_drive.py
"""
from pathlib import Path
import openpyxl

CODERS = ["Leah", "Bridget", "Rachel", "Alia"]
QUESTIONS = ["Q1", "Q2", "Q3", "Q4", "Q5", "Q6"]

HEADER = (
    ["order.num", "num.coders"]
    + [f"q{i}.agree" for i in range(1, 7)]
    + [f"q1.rest.{c}" for c in CODERS]
    + CODERS
    + [
        col for c in CODERS
        for col in [
            f"Q1.{c}", f"Q1.notes.{c}",
            f"Q2.{c}", f"Q2.notes.{c}",
            f"Q3.{c}", f"Q3.notes.{c}",
            f"Q4.{c}", f"Q4.notes.{c}",
            f"Q5.{c}", f"Q5.notes.{c}",
            f"Q6.{c}", f"notes.{c}",
        ]
    ]
)
ORDER_NUM_COL = 0
N_COLS = len(HEADER)


def _blank_row(order: int, n_coders: int) -> list:
    row = [None] * N_COLS
    row[0] = order
    row[1] = n_coders
    # qN.agree defaults True
    for i in range(2, 8):
        row[i] = True
    return row


def _put(row: list, col_name: str, value):
    row[HEADER.index(col_name)] = value


def _coder_block(row: list, coder: str, q1, q2, q3, q4, q5, q6, indicator=True):
    if indicator:
        _put(row, coder, "1")
    _put(row, f"Q1.{coder}", q1)
    _put(row, f"Q2.{coder}", q2)
    _put(row, f"Q3.{coder}", q3)
    _put(row, f"Q4.{coder}", q4)
    _put(row, f"Q5.{coder}", q5)
    _put(row, f"Q6.{coder}", q6)


def _vol_a_data() -> list:
    """20 rows; Leah + Alia code every row."""
    rows = []
    for order in range(1, 21):
        row = _blank_row(order, 2)
        if order <= 15:
            _coder_block(row, "Leah", "no", None, None, None, None, "no")
            _coder_block(row, "Alia", "no", None, None, None, None, "no")
        elif order == 16:
            _coder_block(row, "Leah", "yes", "no", "no", "the public", None, "no")
            _coder_block(row, "Alia", "no",  None, None, None,         None, "no")
            _put(row, "q1.agree", False)
            _put(row, "q4.agree", False)
        elif order == 17:
            _coder_block(row, "Leah", "yes", "no", "no", "the public", "Create or extend comment periods", "yes")
            _coder_block(row, "Alia", "yes", "no", "no", "the public", "Create or extend comment periods", "yes")
        elif order == 18:
            _coder_block(row, "Leah", "yes", "no", "no", "academia, the public", "Advisory board or committee", "yes")
            _coder_block(row, "Alia", "yes", "no", "no", "the public",            "Advisory board or committee", "yes")
            _put(row, "q4.agree", False)
        elif order == 19:
            _coder_block(row, "Leah", "no", None, None, None, None, "no")
            _coder_block(row, "Alia", "no", None, None, None, None, "yes")
            _put(row, "q6.agree", False)
        else:  # 20
            _coder_block(row, "Leah", "no", None, None, None, None, "no")
            _coder_block(row, "Alia", "no", None, None, None, None, "no")
        # q1.rest.<coder> only for Q1 (placeholder "ok" if participated)
        for c in CODERS:
            if row[HEADER.index(c)] == "1":
                _put(row, f"q1.rest.{c}", "ok")
        rows.append(row)
    return rows


def _write_workbook(out: Path, data_rows: list, agreement_sheet_names: list, coder_sheets_filled: dict):
    """Write a workbook with the given agreement sheets and per-coder sheets.

    coder_sheets_filled[coder] = float in [0,1] -- what fraction of rows that coder filled
    in their *coder sheet* (a separate sheet from the agreement sheet).
    """
    wb = openpyxl.Workbook()
    # Remove the default sheet
    default = wb.active
    wb.remove(default)

    # Agreement sheets (all identical content for fixture simplicity)
    for sheet_name in agreement_sheet_names:
        ws = wb.create_sheet(sheet_name)
        ws.append(HEADER)
        for row in data_rows:
            ws.append(row)

    # Per-coder sheets (simplified: order + Q1..Q6 + notes)
    coder_sheet_header = ["Order"] + [
        c
        for q in QUESTIONS
        for c in [f"{q}. (question)", f"{q}(a). notes"]
    ] + ["Notes"]
    for coder in CODERS:
        if coder not in coder_sheets_filled:
            continue
        ws = wb.create_sheet(coder)
        ws.append(coder_sheet_header)
        fill_fraction = coder_sheets_filled[coder]
        n_filled = int(len(data_rows) * fill_fraction)
        for i, row in enumerate(data_rows):
            order = row[ORDER_NUM_COL]
            if i < n_filled:
                # Fill from the data_rows for this coder
                vals = []
                for q in QUESTIONS:
                    vals.append(row[HEADER.index(f"{q}.{coder}")])
                    vals.append(None)
                ws.append([order] + vals + [None])
            else:
                ws.append([order] + [None] * (len(coder_sheet_header) - 1))

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote {out}")


def build_all():
    base = Path(__file__).parent / "fixtures" / "mini_drive_tree"
    data = _vol_a_data()

    # Volume A: complete, 1 agreement sheet
    _write_workbook(
        base / "Volume-A_complete.xlsx",
        data,
        agreement_sheet_names=["33 agreement"],
        coder_sheets_filled={"Leah": 1.0, "Alia": 1.0},
    )

    # Volume B: complete, 2 agreement sheets -- earliest is "38 agreement" (Mar 8)
    _write_workbook(
        base / "Volume-B_two_agreements.xlsx",
        data,
        agreement_sheet_names=["38 agreement", "310 agreement"],
        coder_sheets_filled={"Leah": 1.0, "Alia": 1.0},
    )

    # Volume C: Alia only 30% filled in her coder sheet (the agreement sheet
    # still has the data; the per-coder sheet is what's incomplete)
    _write_workbook(
        base / "Volume-C_incomplete.xlsx",
        data,
        agreement_sheet_names=["33 agreement"],
        coder_sheets_filled={"Leah": 1.0, "Alia": 0.3},
    )

    # Volume D: no agreement sheets at all (only coder sheets)
    _write_workbook(
        base / "Volume-D_no_agreement.xlsx",
        data,
        agreement_sheet_names=[],
        coder_sheets_filled={"Leah": 1.0, "Alia": 1.0},
    )


if __name__ == "__main__":
    build_all()
