import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
import pytest

import assignment_sheet


# ---------- Schema A helpers (single 'Rows' column) ----------

def _build_schema_a(tmp_path, sheet_name, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(["Coder", "Rows", "#Count"])
    for r in rows:
        ws.append(r)
    fpath = tmp_path / f"{sheet_name}.xlsx"
    wb.save(fpath)
    return openpyxl.load_workbook(fpath, read_only=True, data_only=True)


def test_schema_a_simple_two_coder(tmp_path):
    wb = _build_schema_a(tmp_path, "Assignment", [
        ("Leah", "1-200", 200),
        ("Alia", "1-200", 200),
    ])
    result = assignment_sheet.parse(wb)
    assert result["Leah"] == set(range(1, 201))
    assert result["Alia"] == set(range(1, 201))


def test_schema_a_multiple_coders_different_ranges(tmp_path):
    wb = _build_schema_a(tmp_path, "Assignment", [
        ("Alia", "1-400", 400),
        ("Rachel", "1-300", 300),
        ("Brian", "301-600", 300),
        ("Bridget", "401-500", 100),
    ])
    result = assignment_sheet.parse(wb)
    assert result["Alia"] == set(range(1, 401))
    assert result["Rachel"] == set(range(1, 301))
    assert result["Brian"] == set(range(301, 601))
    assert result["Bridget"] == set(range(401, 501))


def test_schema_a_en_dash_in_range(tmp_path):
    wb = _build_schema_a(tmp_path, "Assignment", [
        ("Leah", "1–200", 200),
    ])
    result = assignment_sheet.parse(wb)
    assert result["Leah"] == set(range(1, 201))


def test_schema_a_skips_unknown_coder_names(tmp_path):
    wb = _build_schema_a(tmp_path, "Assignment", [
        ("Leah", "1-100", 100),
        ("SOMETOTAL", "1-100", 100),
        ("", "150-200", 50),
    ])
    result = assignment_sheet.parse(wb)
    assert "Leah" in result
    assert "SOMETOTAL" not in result
    assert "" not in result


def test_schema_a_lowercase_assignments_sheet_name(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "assignments"
    ws.append(["Coder", "Rows", "#Count"])
    ws.append(("Leah", "1-100", 100))
    fpath = tmp_path / "lc.xlsx"
    wb.save(fpath)
    wb2 = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    result = assignment_sheet.parse(wb2)
    assert result["Leah"] == set(range(1, 101))


# ---------- Schema B helpers (numeric start/stop columns, possibly repeated) ----------

def _build_schema_b(tmp_path, sheet_name, rows, n_pairs=2):
    """Build a workbook whose Assignment sheet has the start/stop schema.

    `rows` is a list of (coder, [(start, stop), ...]) tuples. `n_pairs` is
    the maximum number of (start, stop) pairs in any row - defines the
    header width.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    header = ["Coder"] + ["start", "stop"] * n_pairs
    ws.append(header)
    for coder, ranges in rows:
        row = [coder]
        for i in range(n_pairs):
            if i < len(ranges):
                row.extend([ranges[i][0], ranges[i][1]])
            else:
                row.extend([None, None])
        ws.append(row)
    fpath = tmp_path / f"{sheet_name}.xlsx"
    wb.save(fpath)
    return openpyxl.load_workbook(fpath, read_only=True, data_only=True)


def test_schema_b_single_range_per_coder(tmp_path):
    wb = _build_schema_b(tmp_path, "assignments", [
        ("Bridget", [(1, 200)]),
        ("Rachel", [(200, 400)]),
    ], n_pairs=2)
    result = assignment_sheet.parse(wb)
    assert result["Bridget"] == set(range(1, 201))
    assert result["Rachel"] == set(range(200, 401))


def test_schema_b_multiple_disjoint_ranges_per_coder(tmp_path):
    wb = _build_schema_b(tmp_path, "assignments", [
        ("Alia", [(1, 100), (300, 400)]),       # two disjoint ranges
        ("Leah", [(1, 22), (100, 299)]),
        ("Bridget", [(1, 200)]),                # only one range; 2nd pair None
        ("Rachel", [(200, 400)]),
    ], n_pairs=2)
    result = assignment_sheet.parse(wb)
    assert result["Alia"] == set(range(1, 101)) | set(range(300, 401))
    assert result["Leah"] == set(range(1, 23)) | set(range(100, 300))
    assert result["Bridget"] == set(range(1, 201))
    assert result["Rachel"] == set(range(200, 401))


def test_schema_b_three_pair_header(tmp_path):
    """Future-proof: 3 (start, stop) column pairs."""
    wb = _build_schema_b(tmp_path, "Assignment", [
        ("Alia", [(1, 50), (100, 150), (300, 400)]),
    ], n_pairs=3)
    result = assignment_sheet.parse(wb)
    assert result["Alia"] == set(range(1, 51)) | set(range(100, 151)) | set(range(300, 401))


def test_schema_b_skips_unknown_coder_names(tmp_path):
    wb = _build_schema_b(tmp_path, "Assignment", [
        ("Leah", [(1, 100)]),
        ("Unknown", [(1, 100)]),
    ], n_pairs=2)
    result = assignment_sheet.parse(wb)
    assert "Leah" in result
    assert "Unknown" not in result


def test_schema_b_invalid_start_gt_stop_skipped(tmp_path):
    wb = _build_schema_b(tmp_path, "Assignment", [
        ("Leah", [(1, 100), (500, 200)]),   # second range is invalid (start > stop)
    ], n_pairs=2)
    result = assignment_sheet.parse(wb)
    assert result["Leah"] == set(range(1, 101))   # only the valid range


# ---------- Cross-schema and edge cases ----------

def test_no_assignment_sheet_returns_empty(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SomeOtherSheet"
    fpath = tmp_path / "no_assign.xlsx"
    wb.save(fpath)
    wb2 = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    result = assignment_sheet.parse(wb2)
    assert result == {}


def test_empty_assignment_sheet_returns_empty(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assignment"
    # only header, no data rows
    ws.append(["Coder", "Rows", "#Count"])
    fpath = tmp_path / "empty.xlsx"
    wb.save(fpath)
    wb2 = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    result = assignment_sheet.parse(wb2)
    assert result == {}


# ---------- Fallback: derive assignment from agreement-sheet indicator columns ----------

def test_fallback_parses_from_indicator_columns(tmp_path):
    """When no Assignment sheet exists, build coder->orders from per-row indicators."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "109 agreement"
    ws.append([
        "order.num", "num.coders",
        "q1.agree", "q2.agree", "q3.agree", "q4.agree", "q5.agree", "q6.agree",
        "q1.rest.Rachel", "q1.rest.Alia",
        "Rachel", "Alia",
    ])
    # Row 1: both Rachel and Alia
    ws.append([1, 2, True, True, True, True, True, True, "ok", "ok", "1", "1"])
    # Row 2: only Rachel
    ws.append([2, 1, True, True, True, True, True, True, "ok", None, "1", None])
    # Row 3: only Alia
    ws.append([3, 1, True, True, True, True, True, True, None, "ok", None, "1"])
    fpath = tmp_path / "no_assignment.xlsx"
    wb.save(fpath)
    wb2 = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    result = assignment_sheet.parse_from_agreement_indicators(wb2, "109 agreement")
    assert result["Rachel"] == {1, 2}
    assert result["Alia"] == {1, 3}


def test_fallback_ignores_non_coder_indicator_columns(tmp_path):
    """Bare columns named exactly after a known coder are indicators; other
    bare columns must be ignored even if they have '1' values."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "104 agreement"
    ws.append([
        "order.num", "num.coders",
        "q1.agree", "q2.agree", "q3.agree", "q4.agree", "q5.agree", "q6.agree",
        "Leah", "Bridget", "extra_column", "AnotherUnknown",
    ])
    ws.append([1, 2, True, True, True, True, True, True, "1", "1", "1", "1"])
    fpath = tmp_path / "stray_cols.xlsx"
    wb.save(fpath)
    wb2 = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    result = assignment_sheet.parse_from_agreement_indicators(wb2, "104 agreement")
    assert set(result.keys()) == {"Leah", "Bridget"}
    assert "extra_column" not in result
    assert "AnotherUnknown" not in result


def test_fallback_skips_q1_rest_columns(tmp_path):
    """'q1.rest.Rachel' must NOT be treated as a Rachel indicator column."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "33 agreement"
    ws.append([
        "order.num", "num.coders",
        "q1.agree", "q2.agree", "q3.agree", "q4.agree", "q5.agree", "q6.agree",
        "q1.rest.Rachel", "q1.rest.Alia",   # NOT indicators
    ])
    # Even though q1.rest.Rachel is "ok", we shouldn't treat it as the
    # Rachel indicator - there's no bare "Rachel" column here.
    ws.append([1, 2, True, True, True, True, True, True, "ok", "ok"])
    fpath = tmp_path / "rest_only.xlsx"
    wb.save(fpath)
    wb2 = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    result = assignment_sheet.parse_from_agreement_indicators(wb2, "33 agreement")
    # No bare indicator columns => empty result
    assert result == {}


def test_fallback_handles_blank_and_float_indicators(tmp_path):
    """Indicators can be '1', 1, 1.0, or any non-blank truthy value;
    blank/None means the coder did NOT code that row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "agreement"
    ws.append([
        "order.num", "num.coders",
        "q1.agree", "q2.agree", "q3.agree", "q4.agree", "q5.agree", "q6.agree",
        "Leah", "Alia",
    ])
    ws.append([1, 2, True, True, True, True, True, True, 1.0, 1])   # numeric indicators
    ws.append([2, 2, True, True, True, True, True, True, "1", "1"]) # string
    ws.append([3, 1, True, True, True, True, True, True, "1", None])
    ws.append([4, 1, True, True, True, True, True, True, "", "1"])  # empty-string = not participating
    fpath = tmp_path / "varied.xlsx"
    wb.save(fpath)
    wb2 = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    result = assignment_sheet.parse_from_agreement_indicators(wb2, "agreement")
    assert result["Leah"] == {1, 2, 3}
    assert result["Alia"] == {1, 2, 4}


def test_fallback_returns_empty_when_no_indicator_columns(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "104 agreement"
    ws.append([
        "order.num", "num.coders",
        "q1.agree", "q2.agree", "q3.agree", "q4.agree", "q5.agree", "q6.agree",
    ])
    ws.append([1, 2, True, True, True, True, True, True])
    fpath = tmp_path / "bare.xlsx"
    wb.save(fpath)
    wb2 = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    result = assignment_sheet.parse_from_agreement_indicators(wb2, "104 agreement")
    assert result == {}
