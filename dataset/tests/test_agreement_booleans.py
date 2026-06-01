import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
import pytest

import agreement_booleans


def _build_workbook_with_agreement(tmp_path, sheet_name, q_cols, data_rows):
    """Build a workbook whose only sheet is an agreement sheet.

    q_cols is a list like ['q1', 'q2', ..., 'q7'] — header becomes
    'q1.agree', 'q2.agree', etc.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    header = ["order.num", "num.coders"] + [f"{q}.agree" for q in q_cols]
    ws.append(header)
    for r in data_rows:
        ws.append(r)
    fpath = tmp_path / "agree_wb.xlsx"
    wb.save(fpath)
    return openpyxl.load_workbook(fpath, read_only=True, data_only=True)


def test_read_basic_q1_through_q7(tmp_path):
    wb = _build_workbook_with_agreement(
        tmp_path, "104 agreement",
        q_cols=["q1", "q2", "q3", "q4", "q5", "q6", "q7"],
        data_rows=[
            (1, 2, True,  True,  True,  True,  True,  True,  True),
            (2, 2, False, True,  True,  True,  True,  True,  True),
            (3, 2, True,  True,  True,  False, True,  True,  False),
        ],
    )
    result = agreement_booleans.read(wb, "104 agreement")
    assert result[1] == {"Q1": True, "Q2": True, "Q3": True, "Q4": True,
                         "Q5": True, "Q6": True, "Q7": True}
    assert result[2]["Q1"] is False
    assert result[3]["Q4"] is False
    assert result[3]["Q7"] is False


def test_read_only_q1_through_q6(tmp_path):
    """Some volumes (e.g. Vol 134 Part 2) only have Q1-Q6."""
    wb = _build_workbook_with_agreement(
        tmp_path, "104 agreement",
        q_cols=["q1", "q2", "q3", "q4", "q5", "q6"],
        data_rows=[(1, 2, True, True, True, True, True, True)],
    )
    result = agreement_booleans.read(wb, "104 agreement")
    assert set(result[1].keys()) == {"Q1", "Q2", "Q3", "Q4", "Q5", "Q6"}
    assert "Q7" not in result[1]


def test_read_string_true_false_values(tmp_path):
    """openpyxl can yield booleans as the strings 'TRUE'/'FALSE' depending on
    how the workbook was generated. Reader must coerce both."""
    wb = _build_workbook_with_agreement(
        tmp_path, "104 agreement",
        q_cols=["q1", "q2", "q3", "q4", "q5", "q6", "q7"],
        data_rows=[
            (1, 2, "TRUE", "FALSE", "True", "false", None, "", "TRUE"),
        ],
    )
    result = agreement_booleans.read(wb, "104 agreement")
    assert result[1]["Q1"] is True
    assert result[1]["Q2"] is False
    assert result[1]["Q3"] is True
    assert result[1]["Q4"] is False
    assert result[1]["Q5"] is None
    assert result[1]["Q6"] is None
    assert result[1]["Q7"] is True


def test_read_skips_rows_with_blank_order_num(tmp_path):
    wb = _build_workbook_with_agreement(
        tmp_path, "104 agreement",
        q_cols=["q1", "q2", "q3", "q4", "q5", "q6", "q7"],
        data_rows=[
            (1, 2, True, True, True, True, True, True, True),
            (None, None, None, None, None, None, None, None, None),
            (3, 2, False, True, True, True, True, True, True),
        ],
    )
    result = agreement_booleans.read(wb, "104 agreement")
    assert set(result.keys()) == {1, 3}


def test_read_raises_when_sheet_missing_order_num(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "104 agreement"
    ws.append(["something_else", "q1.agree"])
    fpath = tmp_path / "bad.xlsx"
    wb.save(fpath)
    wb2 = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    with pytest.raises(ValueError, match="order.num"):
        agreement_booleans.read(wb2, "104 agreement")


def test_read_raises_when_sheet_has_no_q_agree_columns(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "104 agreement"
    ws.append(["order.num", "some_other_col"])
    ws.append((1, "x"))
    fpath = tmp_path / "bad2.xlsx"
    wb.save(fpath)
    wb2 = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    with pytest.raises(ValueError, match="qN.agree"):
        agreement_booleans.read(wb2, "104 agreement")


def test_read_handles_float_order_num(tmp_path):
    """openpyxl with read_only=True sometimes returns numerics as float."""
    wb = _build_workbook_with_agreement(
        tmp_path, "104 agreement",
        q_cols=["q1", "q2", "q3", "q4", "q5", "q6"],
        data_rows=[(1.0, 2.0, True, True, True, True, True, True)],
    )
    result = agreement_booleans.read(wb, "104 agreement")
    assert 1 in result    # coerced to int
