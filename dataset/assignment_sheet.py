"""Parse the workbook's 'Assignment' sheet into {coder_name: set[order_num]}.

Two schemas are supported in production:

  Schema A: header (Coder, Rows, #Count)
    Rows is a string like "1-400" (single range per coder)

  Schema B: header (Coder, start, stop, start, stop, ...)
    Numeric start/stop columns, repeated to support multiple disjoint
    ranges per coder

The parser inspects the header row to dispatch.
"""
from __future__ import annotations

import re
from typing import Dict, List, Set, Tuple

import openpyxl


KNOWN_CODERS = ("Leah", "Bridget", "Rachel", "Alia", "Brian")
_RANGE_RE = re.compile(r"(\d+)\s*[-–]\s*(\d+)")
_ASSIGNMENT_SHEET_NAMES = ("assignment", "assignments")


def _find_assignment_sheet(wb: openpyxl.Workbook):
    return next(
        (s for s in wb.sheetnames if s.strip().lower() in _ASSIGNMENT_SHEET_NAMES),
        None,
    )


def _start_stop_pairs(header: List) -> List[Tuple[int, int]]:
    """Detect Schema B's (start, stop) column-index pairs from the header.

    Returns a list of (start_col_idx, stop_col_idx) tuples, ordered left-to-right.
    Returns [] if Schema B doesn't apply.
    """
    lower = [str(h).strip().lower() if h is not None else "" for h in header]
    start_cols = [i for i, h in enumerate(lower) if h == "start"]
    stop_cols = [i for i, h in enumerate(lower) if h == "stop"]
    if not start_cols or not stop_cols or len(start_cols) != len(stop_cols):
        return []
    return list(zip(start_cols, stop_cols))


def _parse_schema_a(ws) -> Dict[str, Set[int]]:
    """Schema A: 'Coder | Rows ("1-400") | #Count'."""
    result: Dict[str, Set[int]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        coder = str(row[0]).strip()
        if coder not in KNOWN_CODERS:
            continue
        rng_match = None
        for cell in row[1:]:
            if cell is None:
                continue
            m = _RANGE_RE.search(str(cell))
            if m:
                rng_match = m
                break
        if rng_match is None:
            continue
        start, end = int(rng_match.group(1)), int(rng_match.group(2))
        if start > end:
            continue
        result.setdefault(coder, set()).update(range(start, end + 1))
    return result


def _parse_schema_b(ws, pairs: List[Tuple[int, int]]) -> Dict[str, Set[int]]:
    """Schema B: 'Coder | start | stop | start | stop | ...'."""
    result: Dict[str, Set[int]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        coder = str(row[0]).strip()
        if coder not in KNOWN_CODERS:
            continue
        accum: Set[int] = set()
        for start_idx, stop_idx in pairs:
            if start_idx >= len(row) or stop_idx >= len(row):
                continue
            s, e = row[start_idx], row[stop_idx]
            if s is None or e is None:
                continue
            try:
                start, end = int(s), int(e)
            except (TypeError, ValueError):
                continue
            if start > end:
                continue
            accum.update(range(start, end + 1))
        if accum:
            result[coder] = accum
    return result


def parse(wb: openpyxl.Workbook) -> Dict[str, Set[int]]:
    """Return {coder_name: set of order.num assigned to them}.

    Returns {} if no Assignment sheet exists, or it has no parseable data.
    """
    aname = _find_assignment_sheet(wb)
    if aname is None:
        return {}
    ws = wb[aname]
    header = list(next(ws.iter_rows(values_only=True), []))
    if not header:
        return {}
    pairs = _start_stop_pairs(header)
    if pairs:
        return _parse_schema_b(ws, pairs)
    return _parse_schema_a(ws)
