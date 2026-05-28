"""Per-workbook completeness check: agreement sheet present + per-coder fill threshold."""
from __future__ import annotations

from typing import Dict

import openpyxl


KNOWN_CODERS = ("Leah", "Bridget", "Rachel", "Alia", "Brian")


def _coder_fill_fraction(wb: openpyxl.Workbook, coder: str) -> float:
    if coder not in wb.sheetnames:
        return float("nan")
    ws = wb[coder]
    header = next(ws.iter_rows(values_only=True), None)
    if header is None:
        return 0.0
    # Find Q1 column: header starts with "1." or "Q1." (tolerant)
    q1_idx = None
    for i, h in enumerate(header):
        if not (h and isinstance(h, str)):
            continue
        prefix = h.split(".", 1)[0].strip().upper()
        if prefix in ("1", "Q1"):
            q1_idx = i
            break
    if q1_idx is None:
        return 0.0
    # Find Selection column
    sel_idx = None
    for i, h in enumerate(header):
        if isinstance(h, str) and h.strip().lower() == "selection":
            sel_idx = i
            break
    total = 0
    filled = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or row[0] is None:
            continue
        # Selection-aware denominator
        if sel_idx is not None:
            sel = row[sel_idx] if sel_idx < len(row) else None
            try:
                is_selected = int(sel) == 1
            except (TypeError, ValueError):
                is_selected = (sel == "1")
            if not is_selected:
                continue
        total += 1
        v = row[q1_idx] if q1_idx < len(row) else None
        if v is not None and (not isinstance(v, str) or v.strip() not in ("", "(blank)")):
            filled += 1
    return filled / total if total else 0.0


def check(wb: openpyxl.Workbook, threshold: float = 0.80) -> Dict:
    agreement_sheets = [s for s in wb.sheetnames if "agreement" in s.lower()]
    present_coders = [c for c in KNOWN_CODERS if c in wb.sheetnames]
    fractions = {c: _coder_fill_fraction(wb, c) for c in present_coders}

    if not agreement_sheets:
        return {
            "agreement_sheets": [],
            "coder_fill_fractions": fractions,
            "passes": False,
            "reason": "no_agreement_sheet",
        }
    for c, f in fractions.items():
        if f < threshold:
            return {
                "agreement_sheets": agreement_sheets,
                "coder_fill_fractions": fractions,
                "passes": False,
                "reason": f"coder_{c}_underfilled",
            }
    return {
        "agreement_sheets": agreement_sheets,
        "coder_fill_fractions": fractions,
        "passes": True,
        "reason": "ok",
    }
