"""Read the tracking-status spreadsheet from a sibling folder, parse to a dict.

The tracker's exact schema isn't locked at design time; this module is
tolerant and returns {} on anything it doesn't recognize. The pipeline
then falls back to the heuristic 80%-fill completeness check.
"""
from __future__ import annotations

from typing import Dict, Optional

import openpyxl

from drive_source import DriveSource, WorkbookHandle

KNOWN_CODERS = ("Leah", "Bridget", "Rachel", "Alia", "Brian")
COMPLETE_TOKENS = {"complete", "completed", "done", "finished", "yes"}
CODER_PARTICIPATION_TOKENS = {"yes", "y", "x", "true", "1"}


def _find_tracker_handle(
    src: DriveSource, pattern: str
) -> Optional[WorkbookHandle]:
    pattern_lower = pattern.lower()
    for h in src.list_workbooks():
        if pattern_lower in h.name.lower():
            return h
    return None


def load(
    src: DriveSource, tracker_filename_pattern: str = "tracking status"
) -> Dict[str, Dict]:
    handle = _find_tracker_handle(src, tracker_filename_pattern)
    if handle is None:
        return {}
    try:
        wb = src.open(handle)
    except Exception:
        return {}
    if not wb.sheetnames:
        return {}
    ws = wb[wb.sheetnames[0]]
    header_row = next(ws.iter_rows(values_only=True), None)
    if not header_row:
        return {}
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    try:
        vol_idx = next(i for i, h in enumerate(headers) if h.lower() == "volume")
    except StopIteration:
        return {}
    status_idx = next(
        (i for i, h in enumerate(headers) if "status" in h.lower()), None
    )
    coder_idx = {}
    for c in KNOWN_CODERS:
        for i, h in enumerate(headers):
            if h.lower() == c.lower():
                coder_idx[c] = i
                break

    result: Dict[str, Dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or vol_idx >= len(row) or row[vol_idx] is None:
            continue
        vol = str(row[vol_idx]).strip()
        if not vol:
            continue
        status_raw = ""
        if status_idx is not None and status_idx < len(row) and row[status_idx] is not None:
            status_raw = str(row[status_idx]).strip()
        is_complete = status_raw.lower() in COMPLETE_TOKENS
        expected_coders = []
        for c, ci in coder_idx.items():
            if ci < len(row) and row[ci] is not None:
                cell_text = str(row[ci]).strip().lower()
                if cell_text in CODER_PARTICIPATION_TOKENS:
                    expected_coders.append(c)
        result[vol] = {
            "complete": is_complete,
            "expected_coders": expected_coders if expected_coders else None,
            "raw_status_text": status_raw,
        }
    return result
